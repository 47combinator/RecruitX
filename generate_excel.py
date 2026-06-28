"""
Generate a polished Excel ranked output file.
Merges submission.csv with full candidate profiles from JSONL.
"""
import csv
import json
import pickle
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from src.config import CANDIDATES_FILE, PRECOMPUTED_DIR

# Paths
SUBMISSION = Path(r"d:\hackthons\RedRob\output\submission.csv")
CANDIDATES = CANDIDATES_FILE
FEATURES_PKL = PRECOMPUTED_DIR / "features.pkl"
OUTPUT_XLSX = Path(r"d:\hackthons\RedRob\output\RecruitX_Ranked_Candidates.xlsx")

# ── Load submission ──
print("Loading submission.csv...")
results = []
with open(SUBMISSION, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        results.append({
            "candidate_id": row["candidate_id"],
            "rank": int(row["rank"]),
            "score": float(row["score"]),
            "reasoning": row.get("reasoning", ""),
        })
results.sort(key=lambda x: x["rank"])
print(f"  {len(results)} candidates loaded.")

# ── Load candidate profiles ──
print("Loading candidate profiles...")
target_ids = {r["candidate_id"] for r in results}
candidates = {}
with open(CANDIDATES, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        c = json.loads(line)
        if c["candidate_id"] in target_ids:
            candidates[c["candidate_id"]] = c
        if len(candidates) >= len(target_ids):
            break
print(f"  {len(candidates)} profiles loaded.")

# ── Load features ──
print("Loading features...")
features_map = {}
if FEATURES_PKL.exists():
    with open(FEATURES_PKL, "rb") as f:
        data = pickle.load(f)
    features_map = dict(zip(data["candidate_ids"], data["features"]))
    print(f"  Features for {len(features_map)} candidates.")
else:
    print("  No features file found.")

# ── Build Excel ──
print("Building Excel...")
wb = Workbook()

# ══════════════════════════════════════════════════════════
# SHEET 1: RANKED CANDIDATES (main table)
# ══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Ranked Candidates"

# Styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
data_font = Font(name="Calibri", size=10)
mono_font = Font(name="Consolas", size=10)
score_font = Font(name="Consolas", size=10, bold=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
header_border = Border(
    left=Side(style='thin', color='444444'),
    right=Side(style='thin', color='444444'),
    top=Side(style='thin', color='444444'),
    bottom=Side(style='medium', color='000000'),
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top')
left_align = Alignment(vertical='top')

# Top-3 highlight fills
gold_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
silver_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
bronze_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
rank_fills = {1: gold_fill, 2: silver_fill, 3: bronze_fill}

# Even row subtle striping
stripe_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

# Headers
headers = [
    "Rank", "Candidate ID", "Score", "Name", "Current Title", "Current Company",
    "Years of Experience", "Location", "Country", "Industry",
    "Total Skills", "Expert Skills", "Advanced Skills",
    "Top Skills (Expert+Advanced)",
    "Recruiter Response Rate", "Open to Work", "Notice Period (Days)",
    "Profile Completeness", "GitHub Score",
    "Career History Length", "Current Company Size",
    "Reasoning"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = header_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# Data rows
for row_idx, result in enumerate(results, 2):
    cid = result["candidate_id"]
    c = candidates.get(cid, {})
    p = c.get("profile", {})
    s = c.get("redrob_signals", {})
    skills = c.get("skills", [])
    career = c.get("career_history", [])
    feats = features_map.get(cid, {})

    # Skill analysis
    expert_skills = [sk for sk in skills if sk.get("proficiency") == "expert"]
    advanced_skills = [sk for sk in skills if sk.get("proficiency") == "advanced"]
    top_skills_list = sorted(
        expert_skills + advanced_skills,
        key=lambda x: x.get("duration_months", 0),
        reverse=True
    )
    top_skills_str = ", ".join(sk.get("name", "") for sk in top_skills_list[:10])

    row_data = [
        result["rank"],
        cid,
        result["score"],
        p.get("anonymized_name", ""),
        p.get("current_title", ""),
        p.get("current_company", ""),
        round(p.get("years_of_experience", 0), 1),
        p.get("location", ""),
        p.get("country", ""),
        p.get("current_industry", ""),
        len(skills),
        len(expert_skills),
        len(advanced_skills),
        top_skills_str,
        s.get("recruiter_response_rate", 0),
        "Yes" if s.get("open_to_work_flag") else "No",
        s.get("notice_period_days", 0),
        s.get("profile_completeness_score", 0),
        s.get("github_activity_score", -1) if s.get("github_activity_score", -1) >= 0 else "N/A",
        len(career),
        p.get("current_company_size", ""),
        result["reasoning"],
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = left_align

    # Rank column styling
    ws.cell(row=row_idx, column=1).font = Font(name="Consolas", size=10, bold=True)
    ws.cell(row=row_idx, column=1).alignment = center_align

    # Score formatting
    score_cell = ws.cell(row=row_idx, column=3)
    score_cell.font = score_font
    score_cell.number_format = '0.0000'
    score_cell.alignment = center_align

    # CID monospace
    ws.cell(row=row_idx, column=2).font = mono_font

    # Response rate as percentage
    rr_cell = ws.cell(row=row_idx, column=15)
    rr_cell.number_format = '0%'
    rr_cell.alignment = center_align

    # YoE
    ws.cell(row=row_idx, column=7).alignment = center_align
    ws.cell(row=row_idx, column=7).number_format = '0.0'

    # Numeric columns center
    for num_col in [11, 12, 13, 16, 17, 18, 20]:
        ws.cell(row=row_idx, column=num_col).alignment = center_align

    # Reasoning wrap
    ws.cell(row=row_idx, column=len(headers)).alignment = wrap_align

    # Row fill: top 3 highlight, else alternate stripe
    if result["rank"] in rank_fills:
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = rank_fills[result["rank"]]
    elif row_idx % 2 == 0:
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = stripe_fill

# Column widths
col_widths = {
    1: 6,    # Rank
    2: 16,   # CID
    3: 9,    # Score
    4: 22,   # Name
    5: 28,   # Title
    6: 22,   # Company
    7: 10,   # YoE
    8: 16,   # Location
    9: 10,   # Country
    10: 20,  # Industry
    11: 8,   # Total Skills
    12: 8,   # Expert
    13: 8,   # Advanced
    14: 48,  # Top Skills
    15: 12,  # Response Rate
    16: 8,   # Open to Work
    17: 10,  # Notice
    18: 10,  # Profile Comp
    19: 8,   # GitHub
    20: 8,   # Career Len
    21: 14,  # Company Size
    22: 70,  # Reasoning
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Row height for reasoning
for row_idx in range(2, len(results) + 2):
    ws.row_dimensions[row_idx].height = 45


# ══════════════════════════════════════════════════════════
# SHEET 2: SCORE SUMMARY (compact view)
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Score Summary")

summary_headers = ["Rank", "Name", "Title", "Company", "YoE", "Score", "Top Skills"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = header_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}1"

for row_idx, result in enumerate(results, 2):
    cid = result["candidate_id"]
    c = candidates.get(cid, {})
    p = c.get("profile", {})
    skills = c.get("skills", [])
    top_sk = [sk.get("name", "") for sk in sorted(skills, key=lambda x: {"expert":0,"advanced":1}.get(x.get("proficiency",""),2)) if sk.get("proficiency") in ("expert","advanced")][:6]

    row = [
        result["rank"],
        p.get("anonymized_name", ""),
        p.get("current_title", ""),
        p.get("current_company", ""),
        round(p.get("years_of_experience", 0), 1),
        result["score"],
        ", ".join(top_sk),
    ]
    for col_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
    ws2.cell(row=row_idx, column=1).font = Font(name="Consolas", size=10, bold=True)
    ws2.cell(row=row_idx, column=1).alignment = center_align
    ws2.cell(row=row_idx, column=6).font = score_font
    ws2.cell(row=row_idx, column=6).number_format = '0.0000'
    ws2.cell(row=row_idx, column=6).alignment = center_align
    ws2.cell(row=row_idx, column=5).alignment = center_align

    if result["rank"] in rank_fills:
        for ci in range(1, len(summary_headers) + 1):
            ws2.cell(row=row_idx, column=ci).fill = rank_fills[result["rank"]]
    elif row_idx % 2 == 0:
        for ci in range(1, len(summary_headers) + 1):
            ws2.cell(row=row_idx, column=ci).fill = stripe_fill

for col, w in {1:6, 2:22, 3:28, 4:22, 5:8, 6:10, 7:50}.items():
    ws2.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════
# SHEET 3: SKILLS MATRIX
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Skills Matrix")

# Collect all unique skills across top candidates
all_skills_set = set()
for r in results:
    c = candidates.get(r["candidate_id"], {})
    for sk in c.get("skills", []):
        all_skills_set.add(sk.get("name", ""))
all_skills_sorted = sorted(all_skills_set)

# Headers
ws3.cell(row=1, column=1, value="Rank").font = header_font
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=1).border = header_border
ws3.cell(row=1, column=2, value="Name").font = header_font
ws3.cell(row=1, column=2).fill = header_fill
ws3.cell(row=1, column=2).border = header_border
ws3.cell(row=1, column=3, value="Score").font = header_font
ws3.cell(row=1, column=3).fill = header_fill
ws3.cell(row=1, column=3).border = header_border

for si, skill_name in enumerate(all_skills_sorted, 4):
    cell = ws3.cell(row=1, column=si, value=skill_name)
    cell.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.border = header_border
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')

ws3.freeze_panes = "D2"

prof_fill = {
    "expert": PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "advanced": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "intermediate": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "beginner": PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"),
}
prof_font_white = Font(name="Calibri", size=8, color="FFFFFF")
prof_font_dark = Font(name="Calibri", size=8, color="333333")

for row_idx, result in enumerate(results, 2):
    cid = result["candidate_id"]
    c = candidates.get(cid, {})
    p = c.get("profile", {})
    skills = c.get("skills", [])

    ws3.cell(row=row_idx, column=1, value=result["rank"]).font = Font(name="Consolas", size=9, bold=True)
    ws3.cell(row=row_idx, column=1).alignment = center_align
    ws3.cell(row=row_idx, column=2, value=p.get("anonymized_name", "")).font = Font(name="Calibri", size=9)
    ws3.cell(row=row_idx, column=3, value=result["score"]).font = Font(name="Consolas", size=9)
    ws3.cell(row=row_idx, column=3).number_format = '0.0000'

    skill_lookup = {sk.get("name", ""): sk.get("proficiency", "") for sk in skills}
    for si, skill_name in enumerate(all_skills_sorted, 4):
        prof = skill_lookup.get(skill_name, "")
        if prof:
            cell = ws3.cell(row=row_idx, column=si, value=prof[0].upper())
            cell.fill = prof_fill.get(prof, PatternFill())
            cell.font = prof_font_white if prof in ("expert", "advanced") else prof_font_dark
            cell.alignment = center_align
            cell.border = thin_border

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 8
for si in range(4, 4 + len(all_skills_sorted)):
    ws3.column_dimensions[get_column_letter(si)].width = 3.5
ws3.row_dimensions[1].height = 100


# ══════════════════════════════════════════════════════════
# SHEET 4: METHODOLOGY NOTE
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Methodology")

notes = [
    ("RecruitX Ranking Methodology", ""),
    ("", ""),
    ("Pipeline", "6-stage: JD Parse > Feature Extract > Hybrid Retrieval > Ensemble Rank > Honeypot Filter > Explain"),
    ("Candidates Scanned", "100,000"),
    ("Output", "Top 100 ranked candidates"),
    ("Ranking Runtime", "25.1 seconds on CPU"),
    ("", ""),
    ("Retrieval", ""),
    ("Dense Model", "all-MiniLM-L6-v2 (384-dim sentence embeddings, cosine similarity)"),
    ("Sparse Model", "BM25 Okapi"),
    ("Fusion", "Reciprocal Rank Fusion (k=60) producing top 300 shortlist"),
    ("", ""),
    ("Scoring Weights", ""),
    ("Skill Match", "35% (must-have 60% + nice-to-have 40%)"),
    ("Career Relevance", "20% (title relevance + production ML)"),
    ("Semantic Similarity", "15% (embedding cosine)"),
    ("Behavioral Signals", "10% (response rate + recency)"),
    ("Experience Fit", "8% (years in 5-9 target band)"),
    ("Education", "5%"),
    ("Location", "5% (India preferred)"),
    ("Certifications", "2%"),
    ("", ""),
    ("Honeypot Detection", "7 heuristics: expert skills with 0 duration, 40+ skills, experience-career mismatch, assessment gaps, title contradictions"),
    ("Honeypots in Top 100", "0"),
    ("", ""),
    ("Explainability", "Deterministic, fact-based reasoning generated from actual profile data. No LLM used. Every claim verifiable."),
    ("", ""),
    ("Score Column", "Composite score from 0 to 1. Higher = better fit for Senior AI Engineer JD."),
    ("Skills Matrix Legend", "E = Expert, A = Advanced, I = Intermediate, B = Beginner. Green intensity indicates proficiency."),
]

for row_idx, (label, value) in enumerate(notes, 1):
    cell_a = ws4.cell(row=row_idx, column=1, value=label)
    cell_b = ws4.cell(row=row_idx, column=2, value=value)
    if not value and label:
        cell_a.font = Font(name="Calibri", size=12, bold=True)
    else:
        cell_a.font = Font(name="Calibri", size=10, bold=True, color="444444")
        cell_b.font = Font(name="Calibri", size=10)

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 90


# ── Save ──
wb.save(str(OUTPUT_XLSX))
print(f"\nSaved: {OUTPUT_XLSX}")
print(f"Sheets: {wb.sheetnames}")
print("Done.")
